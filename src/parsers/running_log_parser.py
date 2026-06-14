"""Shared parsing logic for Max's running logs.

Used by:
  - freeze_historical.py  (one-time parse of 2016-2025 xlsx -> historical_daily.csv)
  - build_dataset.py      (hot path: combines historical_daily.csv + current-year xlsx
                           + adjustments -> daily.csv + races.csv)
"""
import csv
import openpyxl
import pandas as pd
import re
import os
from datetime import date, timedelta


# ---------- column schema per year era ----------

SCHEMAS = {
    2017: {"month": 1, "day": 2, "sleep": 3, "miles": 4, "minutes": 5,
           "temp": 7, "weather": 8, "workout": 9, "partners": 10,
           "conditions": None, "wind": None, "time_of_day": None,
           "shoes": None, "location": None, "weight": None,
           "sleep_unit": "hours"},
    2018: {"month": 2, "day": 3, "sleep": 4, "miles": 5, "minutes": 6,
           "temp": 7, "weather": 8, "workout": 9, "partners": 10,
           "conditions": 11, "wind": None, "time_of_day": 12,
           "shoes": None, "location": 13, "weight": None,
           "sleep_unit": "cycles"},
    2020: {"month": 2, "day": 3, "sleep": 4, "miles": 5, "minutes": 6,
           "temp": 7, "weather": 8, "workout": 9, "partners": 10,
           "conditions": 11, "wind": None, "time_of_day": 12,
           "shoes": 13, "location": 14, "weight": None,
           "sleep_unit": "cycles"},
    2022: {"month": 2, "day": 3, "sleep": 4, "miles": 5, "minutes": 6,
           "temp": 7, "weather": 8, "workout": 9, "partners": 10,
           "conditions": 11, "wind": None, "time_of_day": 12,
           "shoes": 13, "location": 14, "weight": 15,
           "sleep_unit": "cycles"},
    2025: {"month": 2, "day": 3, "sleep": 4, "miles": 5, "minutes": 6,
           "temp": 7, "weather": 8, "workout": 9, "partners": 10,
           "conditions": 11, "wind": 12, "time_of_day": 13,
           "shoes": 14, "location": 15, "weight": 16,
           "sleep_unit": "cycles"},
}

# Map each year -> (default_sheet_name, schema_era). Sheet name is a fallback;
# callers can override. Only the schema_era really matters for parsing.
YEAR_SCHEMAS = {
    2017: ("Sheet1", 2017),
    2018: ("Sheet1", 2018),
    2019: ("Sheet1", 2018),
    2020: ("Sheet1", 2020),
    2021: ("Sheet1", 2020),
    2022: ("Sheet1", 2022),
    2023: ("Sheet1", 2022),
    2024: ("Data",   2022),
    2025: ("Data",   2025),
    2026: ("Data",   2025),
    2027: ("Data",   2025),  # assume 2027 uses same schema as 2025 by default
}

# Authoritative annual totals from Lifetime Miles, for validation.
# Extend this when a year finalizes.
AUTHORITATIVE_TOTALS = {
    2016: 2142.6,
    2017: 2205.8,
    2018: 2511.5,
    2019: 3183.5,
    2020: 3200.0,
    2021: 2030.3,
    2022: 2644.0,
    2023: 3218.9,
    2024: 3137.0,
    2025: 3232.3,
}

DAILY_COLUMNS = [
    "date", "year", "month", "day_of_year", "dow",
    "miles", "minutes", "pace_sec_per_mi",
    "temp_c", "sleep_cycles",
    "weather", "conditions", "wind", "wind_ms", "humidity_pct", "time_of_day",
    "shoes", "location", "weight_lbs", "surface",
    "partners", "workout_raw",
    "run_type",
    "recovery_pace_sec_per_mi",
    "quality_distance_m", "quality_pace_sec_per_mi", "quality_segment_type",
    "num_races",
    "schema_year_era", "source_file",
]


# ---------- small utilities ----------

def _safe_float(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_time_to_seconds(time_str):
    """Parse 'SS.cc', 'M:SS(.cc)?', or 'H:MM:SS(.cc)?' -> float seconds."""
    if time_str is None:
        return None
    s = str(time_str).strip()
    try:
        parts = s.split(":")
        if len(parts) == 1:
            return float(parts[0])
        if len(parts) == 2:
            mins, secs = parts
            return int(mins) * 60 + float(secs)
        if len(parts) == 3:
            hrs, mins, secs = parts
            return int(hrs) * 3600 + int(mins) * 60 + float(secs)
    except (ValueError, AttributeError):
        return None
    return None


# ---------- surface classification (location-based, priority 2) ----------

def surface_from_location(location_str):
    """Classify running surface from the location string.

    Returns 'Track' if the location contains any of the configured track
    keywords, else None. `None` means "no opinion" — the caller decides
    whether to apply further rules or fall back to 'Unknown'.
    """
    if location_str is None:
        return None
    s = str(location_str).strip().lower()
    if s == "":
        return None
    for kw in ("shoreline", "track", "rose park", "niwot", "outdoor rec"):
        if kw in s:
            return "Track"
    return None


# ---------- workout string classification ----------

def classify_run_type(workout_str, year):
    """Classify a workout string into a single run-type tag.

    Precedence mirrors the Google Sheets conditional formatting (first match wins).
    In 2016 `Nr@` means race; in 2017+ `Nr@` means reps and races use `race@`.

    Returns None when unclassifiable — callers should apply their own
    mileage-based fallback (0 miles -> rest, >0 miles -> recovery).
    """
    if workout_str is None:
        return None
    s = str(workout_str).strip()
    if s == "":
        return None
    sl = s.lower()
    if sl == "rest":
        return "rest"
    # Pure cross-training days (no running component)
    if re.fullmatch(r"(lifting|weights|strength|yoga|core|swim|bike)", sl):
        return "cross_train"
    # Strip common strength/cross-training prefixes so "lifting, rec@6:13" classifies as recovery
    s_run = re.sub(r"^(lifting|weights|strength|yoga|core|swim|bike)[,\s]+", "", sl).strip()
    # 2016-only: "rec@<pace> long" was Max's notation for subjective long runs
    # before the convention switched to "long@<pace>" in 2017+.
    if year == 2016 and re.match(r"^rec@", s_run) and re.search(r"\blong\s*$", s_run):
        return "long"
    if re.match(r"^rec(\b|@|\s|/)", s_run): return "recovery"
    if re.match(r"^long(\b|@|\s|/)", s_run): return "long"
    if re.match(r"^trail(\b|@|\s|/)", s_run): return "trail"
    # Race detection (year-dependent)
    if year == 2016:
        # In 2016, a distance followed by 'r@' = race (e.g., '1600r@4:57').
        # Also allow 'Nr <time>' with a space separator instead of @.
        if re.search(r"\d\s*r@", sl) or "race@" in sl:
            return "race"
        if re.search(r"\d\s*r\s+\d+[:.]?\d", sl):
            return "race"
    else:
        if "race@" in sl:
            return "race"
    # Hills — 2016 sometimes used `hc/` and `hr/` where later years use `hc-`/`hr-`
    if re.search(r"hc[-/]", sl): return "hill_cont"
    if re.search(r"hr[-/]", sl): return "hill_rep"
    # Quality types — match <digit><letter> followed by @ OR /
    if re.search(r"\dt[@/]", sl): return "tempo"
    if re.search(r"\di[@/]", sl): return "interval"
    if re.search(r"\df[@/]", sl): return "fartlek"
    if year >= 2017 and re.search(r"\dr[@/]", sl): return "rep"
    if year == 2016 and re.search(r"\drep[@/]", sl): return "rep"
    return None


def extract_recovery_pace(workout_str):
    """For rec/long/trail cells, extract the @M:SS pace (per mile, in seconds)."""
    if not workout_str:
        return None
    m = re.match(r"^(rec|long|trail)@([\d:.]+)", str(workout_str).strip().lower())
    if m:
        return parse_time_to_seconds(m.group(2))
    return None


def extract_quality_segment(workout_str, year):
    """For tempo/interval/rep/fartlek workouts, extract first quality segment.

    Returns (distance_m, pace_sec_per_mi, segment_type) or (None, None, None).
    """
    if not workout_str:
        return (None, None, None)
    s = str(workout_str)
    if year >= 2017:
        pat = re.compile(r"(\d+)\s*([tirf])@([\d:.]+)")
        type_map = {"t": "tempo", "i": "interval", "r": "rep", "f": "fartlek"}
    else:  # 2016: use t/i/f + explicit 'rep'
        pat = re.compile(r"(\d+)\s*([tif])@([\d:.]+)|(\d+)\s*rep@([\d:.]+)")
        type_map = {"t": "tempo", "i": "interval", "f": "fartlek"}
    m = pat.search(s)
    if not m:
        return (None, None, None)
    if year >= 2017:
        dist = int(m.group(1))
        t = type_map[m.group(2)]
        pace = parse_time_to_seconds(m.group(3))
    else:
        if m.group(1):
            dist = int(m.group(1))
            t = type_map[m.group(2)]
            pace = parse_time_to_seconds(m.group(3))
        else:
            dist = int(m.group(4))
            t = "rep"
            pace = parse_time_to_seconds(m.group(5))
    return (dist, pace, t)


def extract_races(workout_str, year):
    """Extract all race segments in order. Returns list of
    {distance_m, time_sec, event}.

    Event-name attribution (two sources, with adjacent winning over trailing):
      1. Adjacent: `race@time [Event]` — bracket immediately follows (with optional
         whitespace) the race time. Applies only to that race.
      2. Trailing: `... [Event]` at end of the entire workout string. Applies
         to every race in the workout that doesn't already have an adjacent
         bracket. This is the common case for 2017 HS-era logs where events
         are recorded once at the end of the line.

    Handles:
      - 2017+ `N race@time` syntax, with optional inline event `[Event Name]`
        (e.g. `5000 race@16:00 [Run for the Pies]`)
      - 2016 `Nr@time` and legacy `Nr <time>` (space separator)
      - 4x400 relay legs where Max runs legs 1 and 4 ("2x400 race@a/b" or "2x400@a/b"),
        only on days that already had at least one other race, and with the first
        leg deduped against any bare 400m entry the greedy race@ regex picked up.
    """
    if not workout_str:
        return []
    s = str(workout_str)

    # Trailing event: a `[Event]` at the very end of the string (optional
    # trailing whitespace). This is the fallback event for races that don't
    # have an adjacent bracket.
    trailing_m = re.search(r"\[([^\]]*)\]\s*$", s)
    trailing_event = trailing_m.group(1).strip() if trailing_m else None
    if trailing_event == "":
        trailing_event = None

    candidates = []  # (position, distance_m, time_sec, event_or_None)
    # Adjacent bracket may be separated by whitespace only. A comma or any
    # other content between race@time and `[` means the bracket is NOT
    # adjacent (it's the trailing one, handled below).
    _BRACKET = r"(?:\s*\[([^\]]*)\])?"

    def _pick_event(adj):
        """Adjacent bracket wins; else fall back to trailing."""
        if adj is None:
            return trailing_event
        adj = adj.strip()
        return adj if adj else trailing_event

    if year == 2016:
        for m in re.finditer(r"(\d+)\s*r@([\d:.]+)" + _BRACKET, s):
            candidates.append((m.start(), int(m.group(1)),
                               parse_time_to_seconds(m.group(2)),
                               _pick_event(m.group(3))))
        for m in re.finditer(r"(\d+)\s*r\s+(\d+[:.]?\d[\d:.]*)" + _BRACKET, s.lower()):
            dist = int(m.group(1))
            time_sec = parse_time_to_seconds(m.group(2))
            already = any(d == dist and t == time_sec for _, d, t, _e in candidates)
            if not already:
                candidates.append((m.start(), dist, time_sec, _pick_event(m.group(3))))
    else:
        for m in re.finditer(r"(\d+)\s*race@([\d:.]+)" + _BRACKET, s):
            candidates.append((m.start(), int(m.group(1)),
                               parse_time_to_seconds(m.group(2)),
                               _pick_event(m.group(3))))

    # Relay legs only count if there's already at least one race on this day
    if candidates and year != 2016:
        for m in re.finditer(r"2x400(?:\s*race)?@([\d:.]+)/([\d:.]+)" + _BRACKET, s):
            t1 = parse_time_to_seconds(m.group(1))
            t2 = parse_time_to_seconds(m.group(2))
            event = _pick_event(m.group(3))
            leg_times = [t for t in (t1, t2) if t is not None]

            def _is_dup_leg(c, _leg_times=leg_times):
                _pos, d, t_c, _e = c
                if d != 400 or t_c is None:
                    return False
                return any(abs(t_c - lt) < 0.05 for lt in _leg_times)

            candidates = [c for c in candidates if not _is_dup_leg(c)]
            candidates.append((m.start(), 400, t1, event))
            candidates.append((m.start() + 1, 400, t2, event))

    candidates.sort(key=lambda x: x[0])
    return [{"distance_m": d, "time_sec": t, "event": e}
            for _, d, t, e in candidates]


def split_2016_notes(notes_str):
    """Split 2016's combined workout+partners cell into (workout, partners).

    Formats: '<workout> solo' | '<workout> w/ <partners>' | '<workout>'.

    The trailing 'long' marker (2016-only) is preserved on the workout string
    so classify_run_type can distinguish subjective long runs from recovery.

    A trailing `[Event]` is preserved on the workout string so extract_races
    can attribute the event to its race(s). Without this step the greedy
    partner-splitting regex would swallow the bracket into `partners`.
    """
    if not notes_str:
        return (None, None)
    s = str(notes_str).strip()
    if s == "":
        return (None, None)
    if s.lower() == "rest":
        return ("rest", None)

    # Peel off a trailing event bracket before doing the partner split, then
    # stitch it back onto the workout at the end.
    trailing = ""
    br_m = re.search(r"\s*(\[[^\]]*\])\s*$", s)
    if br_m:
        trailing = " " + br_m.group(1)
        s = s[:br_m.start()].rstrip()

    m = re.search(r"\s+w/\s+(.+)$", s)
    if m:
        partners = m.group(1).strip()
        workout = s[:m.start()].strip()
        return (workout + trailing, partners)
    m = re.search(r"\s+solo\s*$", s, re.I)
    if m:
        workout = s[:m.start()].strip()
        return (workout + trailing, "solo")
    return (s + trailing, None)


# ---------- 2016-17 location inference ----------
#
# The 2016 and 2017 schemas have no location column, so daily rows from
# those years arrive at _derive_daily_row with location=None.
#
# This module covers the *hill-loop* part of synthesis: parse the loop
# abbreviation from workout_raw (e.g. '23hc-8x pwr1' -> 'pwr1') and look
# it up in the hills sheet to get the canonical route name ('powerline
# west'). That signal is load-bearing for recovery-analysis route betas.
#
# Date-range location rules (Nashville/Geneva trips, 'education hill'
# default) used to live here too but were migrated to the snapshot's
# `historical` section in 2026 — see build_dataset.py's apply-historical
# step. Non-hill 2016-17 rows now arrive at daily.csv with location=None
# and get filled in by the historical override at build time.

# Loop name appears after `hc` or `hr` plus optional dash/slash modifiers
# and an optional Nx repcount. Handles all 2016-17 shapes:
#   23hc-8x pwr1     | 1hr-10x pwr2/7x10sp     | 1:30hr-9x evst/8x10sp
#   26hc/rep pwr2    | 25hc/rep-6x fm2/6x10sp  | 30hc-10x fm1, 11j
_HILL_LOOP_RX = re.compile(r"h[cr][^\s]*\s+([a-z0-9]+)")


def _extract_hill_loop(workout_raw):
    """Return the hill loop abbreviation parsed from workout_raw, or None."""
    if not workout_raw:
        return None
    m = _HILL_LOOP_RX.search(str(workout_raw).lower())
    return m.group(1) if m else None


def ingest_hills_from_df(hills_df):
    """Convert a hills DataFrame into the {abbrev_lower: location_lower} dict.

    The "abbrev" column may contain comma-separated synonyms (e.g.
    "evst, ev" or "fm1, fm2") — each is registered as a separate key
    pointing to the same location.
    """
    if hills_df is None or len(hills_df) == 0:
        return {}
    required = {"abbrev", "location"}
    missing = required - set(hills_df.columns)
    if missing:
        print(f"[hills] missing columns: {missing} — skipping")
        return {}
    lookup = {}
    for _, row in hills_df.iterrows():
        abbrev_raw = row["abbrev"]
        loc_raw = row["location"]
        if abbrev_raw in (None, "") or loc_raw in (None, ""):
            continue
        loc = str(loc_raw).strip().lower()
        for tok in str(abbrev_raw).split(","):
            tok = tok.strip().lower()
            if tok:
                lookup[tok] = loc
    return lookup


def infer_2016_2017_location(workout_raw, run_type, hill_lookup):
    """Synthesize a log_location for 2016-17 hill workouts.

    Returns the canonical route name (e.g. 'powerline west') for hill_cont
    / hill_rep workouts when the loop abbreviation in workout_raw matches
    the hills sheet; returns None otherwise. Non-hill 2016-17 rows are
    expected to fall through to the historical-section override in
    build_dataset.py for both city_state and location.
    """
    if run_type in ("hill_cont", "hill_rep"):
        loop = _extract_hill_loop(workout_raw)
        if loop and hill_lookup and loop in hill_lookup:
            return hill_lookup[loop]
    return None


# ---------- xlsx ingestion (freeze path only) ----------

def _derive_daily_row(dt, year, schema_key, source_file,
                      miles, minutes, temp, sleep_cycles,
                      weather, workout, partners, conditions, wind,
                      time_of_day, shoes, location, weight,
                      wind_ms=None, humidity_pct=None,
                      hill_lookup=None):
    """Assemble one daily row dict from already-normalized cell values.

    `hill_lookup` is consulted only for 2016 and 2017 hill workouts when
    `location` is None — see infer_2016_2017_location. Non-hill 2016-17
    rows leave with location=None and get filled in at build time via the
    snapshot's `historical` section.

    Note: callers prune zero-mile rows before invoking this helper (see
    ingest_2016 / ingest_year_standard / ingest_year_standard_csv), so
    every row reaching this point represents an actual run.
    """
    rtype = classify_run_type(workout, year)
    if rtype is None:
        rtype = "recovery"

    # 2016-17 hill-loop synthesis: only acts when caller hasn't already set
    # one (the 2016/2017 xlsx schemas don't have a location column, so
    # `location` arrives as None for these years). Non-hill rows leave
    # location=None and get filled by build_dataset's historical override.
    if location is None and year in (2016, 2017):
        location = infer_2016_2017_location(workout, rtype, hill_lookup or {})

    pace_sec_per_mi = None
    if miles and miles > 0 and minutes:
        pace_sec_per_mi = (minutes * 60) / miles

    rec_pace = extract_recovery_pace(workout)
    q_dist, q_pace, q_type = extract_quality_segment(workout, year)
    races = extract_races(workout, year)

    return {
        "date": dt,
        "year": dt.year,
        "month": dt.month,
        "day_of_year": dt.timetuple().tm_yday,
        "dow": dt.isoweekday(),
        "miles": miles,
        "minutes": minutes,
        "pace_sec_per_mi": pace_sec_per_mi,
        "temp_c": temp,
        "sleep_cycles": sleep_cycles,
        "weather": weather,
        "conditions": conditions,
        "wind": wind,
        "wind_ms": wind_ms,
        "humidity_pct": humidity_pct,
        "time_of_day": time_of_day,
        "shoes": shoes,
        "location": location,
        "weight_lbs": weight,
        "surface": surface_from_location(location) or "Unknown",
        "partners": partners,
        "workout_raw": workout,
        "run_type": rtype,
        "recovery_pace_sec_per_mi": rec_pace,
        "quality_distance_m": q_dist,
        "quality_pace_sec_per_mi": q_pace,
        "quality_segment_type": q_type,
        "num_races": len(races),
        "schema_year_era": schema_key,
        "source_file": source_file,
    }


def ingest_year_standard(path, year, sheet_name=None, hill_lookup=None):
    """Ingest one standard-format year (2017+) from a row-indexed daily log.

    Dates are derived by row position (row 3 = day 1 of year) — robust to
    leap days, annotation rows, and blank months.

    `hill_lookup` is forwarded to _derive_daily_row; see
    infer_2016_2017_location for how it's used.
    """
    if sheet_name is None:
        sheet_name, _ = YEAR_SCHEMAS.get(year, ("Data", 2025))
    _, schema_key = YEAR_SCHEMAS.get(year, (None, 2025))
    sch = SCHEMAS[schema_key]
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        # Fall back to the first sheet; print a warning for visibility
        alt = wb.sheetnames[0]
        print(f"[ingest] {path}: sheet '{sheet_name}' not found, using '{alt}'")
        sheet_name = alt
    ws = wb[sheet_name]
    source_file = os.path.basename(path)

    rows = []
    max_doy = 366 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 365
    for r in range(3, 3 + max_doy):
        miles_v = ws.cell(row=r, column=sch["miles"]).value
        minutes_v = ws.cell(row=r, column=sch["minutes"]).value
        workout_v = ws.cell(row=r, column=sch["workout"]).value
        sleep_v = ws.cell(row=r, column=sch["sleep"]).value
        if (miles_v in (None, "") and minutes_v in (None, "")
                and workout_v in (None, "") and sleep_v in (None, "")):
            continue
        # Prune zero-mile days (rest, cross-training, blank-mile entries).
        # Daily.csv is a running-only dataset; absent rows = days not run.
        miles_check = _safe_float(miles_v)
        if miles_check is None or miles_check == 0:
            continue

        doy = r - 2
        dt = date(year, 1, 1) + timedelta(days=doy - 1)

        def get(field):
            col = sch.get(field)
            if col is None:
                return None
            v = ws.cell(row=r, column=col).value
            return None if v in (None, "") else v

        sleep_num = _safe_float(sleep_v)
        if sleep_num is None:
            sleep_cycles = None
        else:
            sleep_cycles = sleep_num / 1.5 if sch["sleep_unit"] == "hours" else sleep_num

        miles = _safe_float(miles_v)
        minutes = _safe_float(minutes_v)
        temp = _safe_float(get("temp"))
        weather = str(get("weather")).lower() if get("weather") else None
        workout = str(workout_v).strip() if workout_v else None
        partners = get("partners")
        partners = str(partners).strip() if partners else None
        conditions = get("conditions")
        conditions = str(conditions).strip().lower() if conditions else None
        wind = get("wind")
        wind = str(wind).strip().lower() if wind else None
        time_of_day = get("time_of_day")
        time_of_day = str(time_of_day).strip().lower() if time_of_day else None
        shoes = get("shoes")
        shoes = str(shoes).strip() if shoes else None
        location = get("location")
        location = str(location).strip().lower() if location else None
        weight = _safe_float(get("weight"))

        rows.append(_derive_daily_row(
            dt, year, schema_key, source_file,
            miles, minutes, temp, sleep_cycles,
            weather, workout, partners, conditions, wind,
            time_of_day, shoes, location, weight,
            hill_lookup=hill_lookup,
        ))
    return rows


def ingest_2016(path, hill_lookup=None):
    """Ingest 2016 from the monthly-sheet format (month001..month012).

    `hill_lookup` is forwarded to _derive_daily_row; see
    infer_2016_2017_location for how it's used.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    source_file = os.path.basename(path)
    rows = []
    for m in range(1, 13):
        sn = f"month{m:03d}"
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(4, 40):
            day_v = ws.cell(row=r, column=1).value
            if not isinstance(day_v, (int, float)):
                continue
            day_n = int(day_v)
            if not (1 <= day_n <= 31):
                continue
            try:
                dt = date(2016, m, day_n)
            except ValueError:
                continue
            miles_v = ws.cell(row=r, column=3).value
            minutes_v = ws.cell(row=r, column=4).value
            temp_v = ws.cell(row=r, column=6).value
            weather_v = ws.cell(row=r, column=7).value
            notes_v = ws.cell(row=r, column=8).value
            sleep_v = ws.cell(row=r, column=2).value

            if miles_v in (None, "") and notes_v in (None, ""):
                continue

            miles = _safe_float(miles_v)
            # Prune zero-mile days (rest, cross-training, blank-mile entries).
            # Daily.csv is a running-only dataset; absent rows = days not run.
            if miles is None or miles == 0:
                continue
            minutes = _safe_float(minutes_v)
            temp = _safe_float(temp_v)
            weather = str(weather_v).lower() if weather_v else None
            workout, partners = split_2016_notes(notes_v)
            sleep_num = _safe_float(sleep_v)
            sleep_cycles = sleep_num / 1.5 if sleep_num is not None else None

            rows.append(_derive_daily_row(
                dt, 2016, 2016, source_file,
                miles, minutes, temp, sleep_cycles,
                weather, workout, partners, None, None,
                None, None, None, None,
                hill_lookup=hill_lookup,
            ))
    return rows


# ---------- CSV ingestion (hot-path, post-2025 schema) ----------

# Columns expected in the per-year CSV, in order:
#   date (ISO YYYY-MM-DD), sleep_cycles, miles, minutes, temp_c, weather,
#   workout_raw, partners, conditions, wind, time_of_day, shoes, location,
#   weight_lbs
#
# Date-position indexing is NOT used here — the CSV has one row per logged
# day and every row carries its own ISO date, so rest days are simply absent.
# Days beyond the last logged entry are implicitly absent too, which matches
# how build_dataset treats the current-year partial.
CSV_LOG_COLUMNS = [
    "date", "sleep_cycles", "miles", "minutes", "temp_c", "weather",
    "workout_raw", "partners", "conditions", "wind", "time_of_day",
    "shoes", "location", "weight_lbs",
]


def ingest_year_standard_csv(df_or_path, year, source_file=None):
    """Ingest a per-year CSV (or already-loaded DataFrame) for years 2017+.

    This is the hot-path current-year ingest, replacing xlsx for builds. The
    schema is just what the parser needs — no row/column padding, no presets,
    no synthesis tables. Every row has an ISO date; rest days are implicit
    (absent rows).

    Zero-mile rows (rest, cross-training, or blank-miles entries) are pruned
    so daily.csv stays a running-only dataset; consistent with ingest_2016 /
    ingest_year_standard's freeze-time behavior.

    Pass either a filesystem path or a pandas DataFrame. Returns a list of
    daily row dicts matching DAILY_COLUMNS.
    """
    _, schema_key = YEAR_SCHEMAS.get(year, (None, 2025))

    if hasattr(df_or_path, "columns"):
        df = df_or_path
        if source_file is None:
            source_file = f"snapshot:current_log:{year}"
    else:
        df = pd.read_csv(df_or_path)
        if source_file is None:
            source_file = os.path.basename(df_or_path)

    missing = [c for c in CSV_LOG_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"current-year CSV missing columns: {missing}")

    rows = []
    for _, r in df.iterrows():
        date_v = r["date"]
        if pd.isna(date_v) or str(date_v).strip() == "":
            continue
        if hasattr(date_v, "hour"):
            dt = date_v.date()
        elif hasattr(date_v, "year"):
            dt = date_v
        else:
            try:
                dt = pd.Timestamp(date_v).date()
            except (ValueError, TypeError):
                continue
        if dt.year != year:
            # Different year than the one we're ingesting — skip rather than
            # silently mis-assign.
            continue

        def _str(col):
            v = r.get(col)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return None
            s = str(v).strip()
            return s if s else None

        miles = _safe_float(r.get("miles"))
        # Prune zero-mile days (rest, cross-training, blank-mile entries).
        # Daily.csv is a running-only dataset; absent rows = days not run.
        if miles is None or miles == 0:
            continue
        minutes = _safe_float(r.get("minutes"))
        temp = _safe_float(r.get("temp_c"))
        sleep_cycles = _safe_float(r.get("sleep_cycles"))
        weight = _safe_float(r.get("weight_lbs"))
        weather = (_str("weather") or "").lower() or None
        workout = _str("workout_raw")
        partners = _str("partners")
        conditions = (_str("conditions") or "").lower() or None
        wind = (_str("wind") or "").lower() or None
        # Numeric watch weather (current_log from build_current_log). Optional:
        # absent in hand-logged drive snapshots, so default to None.
        wind_ms = _safe_float(r.get("wind_ms"))
        humidity_pct = _safe_float(r.get("humidity_pct"))
        time_of_day = (_str("time_of_day") or "").lower() or None
        shoes = _str("shoes")
        location = (_str("location") or "").lower() or None

        rows.append(_derive_daily_row(
            dt, year, schema_key, source_file,
            miles, minutes, temp, sleep_cycles,
            weather, workout, partners, conditions, wind,
            time_of_day, shoes, location, weight,
            wind_ms=wind_ms, humidity_pct=humidity_pct,
        ))
    return rows


# ---------- race segment construction ----------

# Columns of a race-segment dict (the order build_race_segments emits them).
# Used to construct a well-formed empty races frame when a profile has no
# races at all, so downstream apply/adjustment/summary steps no-op instead of
# failing on a column-less DataFrame.
RACE_SEGMENT_COLUMNS = [
    "date", "year", "race_seq", "fatigued", "distance_m", "time_sec",
    "pace_sec_per_mi", "location", "temp_c", "weather", "conditions", "shoes",
    "surface", "event", "note", "surface_source", "workout_raw", "source_file",
]


def build_race_segments(daily_rows):
    """Given a list of daily row dicts, return a list of race segment dicts.

    Multi-race days get race_seq 1, 2, 3... with fatigued=True for seq>1.
    Surface here is the *location-based* classification only; apply_race_rules
    fills in the rest as a post-processing pass.
    """
    out = []
    # Allow dicts (from parsing) or pandas Series (from loaded CSV) — access uniformly
    def g(row, key):
        return row[key] if not hasattr(row, "get") else row.get(key)

    for row in daily_rows:
        races = extract_races(g(row, "workout_raw"), g(row, "year"))
        for i, race in enumerate(races):
            dist_m = race["distance_m"]
            time_sec = race["time_sec"]
            miles_eq = dist_m / 1609.344
            pace = time_sec / miles_eq if (time_sec and miles_eq > 0) else None
            loc_surface = surface_from_location(g(row, "location"))
            out.append({
                "date": g(row, "date"),
                "year": g(row, "year"),
                "race_seq": i + 1,
                "fatigued": i > 0,
                "distance_m": dist_m,
                "time_sec": time_sec,
                "pace_sec_per_mi": pace,
                "location": g(row, "location"),
                "temp_c": g(row, "temp_c"),
                "weather": g(row, "weather"),
                "conditions": g(row, "conditions"),
                "shoes": g(row, "shoes"),
                "surface": loc_surface,
                "event": race.get("event"),
                "note": None,
                "surface_source": "location" if loc_surface else None,
                "workout_raw": g(row, "workout_raw"),
                "source_file": g(row, "source_file"),
            })
    return out


# ---------- race surface rules (priority 3, distance + date) ----------

_HM_LO, _HM_HI = 21097.5 * 0.92, 21097.5 * 1.08
_MAR_LO, _MAR_HI = 42195 * 0.92, 42195 * 1.08
_5K_LO, _5K_HI = 5000 * 0.92, 5000 * 1.08
_10K_LO, _10K_HI = 10000 * 0.92, 10000 * 1.08


def apply_race_rules(row):
    """Fill surface for rows where priority 1-2 left it blank.

    Order matters — specific rules before fall-throughs so defaults can't
    swallow real XC races.
    """
    if row["surface"] in ("Road", "Track", "XC"):
        return row["surface"]
    dist = row["distance_m"]
    if dist is None or pd.isna(dist):
        return "Unknown"
    d = row["date"]
    # Handle both datetime.date and pd.Timestamp
    if hasattr(d, "year"):
        y, mo, dy = d.year, d.month, d.day
    else:
        return "Unknown"

    # 5K-specific XC-season window for Max's 2016-17 HS XC races
    if _5K_LO <= dist <= _5K_HI:
        if (y == 2016 and mo >= 8) or (y == 2017 and (mo < 6 or (mo == 6 and dy == 1))):
            return "XC"

    if _HM_LO <= dist <= _HM_HI or _MAR_LO <= dist <= _MAR_HI:
        return "Road"
    if _5K_LO <= dist <= _5K_HI:
        return "Road"
    if _10K_LO <= dist <= _10K_HI:
        return "Road"
    if dist < 3218:
        return "Track"
    # 3218m 2-mile stays Unknown — ambiguous track vs road pre-location-data
    return "Unknown"


def set_race_surface_source(row):
    if row.get("surface_source") in ("routes", "location", "addition", "adjustment"):
        return row["surface_source"]
    if row["surface"] in ("Road", "Track", "XC", "Downhill"):
        return "rule"
    return "unknown"


# ---------- adjustments ----------

_NUMERIC_FIELDS = {"distance_m", "time_sec", "pace_sec_per_mi", "temp_c",
                   "race_seq"}


def _coerce_value(field, value):
    if value is None:
        return None
    if field in _NUMERIC_FIELDS:
        try:
            return float(value)
        except (ValueError, TypeError):
            return value
    if isinstance(value, str):
        return value.strip()
    return value


def apply_adjustments_from_df(races_df, changes_df):
    """Apply manual corrections from a changes DataFrame.

    Expected columns: date | race_seq | field | value | note
    Matches by (date, race_seq). Loud on no-match / ambiguous-match / unknown-field.
    When field == 'surface', also sets surface_source = 'adjustment'.

    Returns (races_df, applied_count, failed_count).
    """
    if changes_df is None or len(changes_df) == 0:
        return races_df, 0, 0

    required = {"date", "race_seq", "field", "value"}
    missing = required - set(changes_df.columns)
    if missing:
        print(f"[adjustments] missing columns: {missing} — skipping")
        return races_df, 0, 0

    applied = 0
    failed = 0
    races_dates = pd.to_datetime(races_df["date"]).dt.date

    for r_idx, row in changes_df.iterrows():
        date_v = row["date"]
        seq_v = row["race_seq"]
        field = row["field"]
        value = row["value"]
        note = row.get("note") if "note" in changes_df.columns else None

        if pd.isna(date_v) or not field or (isinstance(field, float) and pd.isna(field)):
            continue

        if hasattr(date_v, "hour"):
            dt = date_v.date()
        elif hasattr(date_v, "year"):
            dt = date_v
        else:
            dt = pd.Timestamp(date_v).date()
        try:
            seq = int(seq_v) if not pd.isna(seq_v) else 1
        except (ValueError, TypeError):
            print(f"[adjustments] row {r_idx}: bad race_seq {seq_v!r} — skipping")
            failed += 1
            continue
        field = str(field).strip()

        mask = (races_dates == dt) & (races_df["race_seq"] == seq)
        nmatch = int(mask.sum())
        if nmatch == 0:
            print(f"[adjustments] NO MATCH  row {r_idx}  {dt} seq={seq}  {field}={value!r}")
            failed += 1
            continue
        if nmatch > 1:
            print(f"[adjustments] AMBIGUOUS row {r_idx}  {dt} seq={seq}  matched {nmatch}")
            failed += 1
            continue
        if field not in races_df.columns:
            print(f"[adjustments] UNKNOWN FIELD  row {r_idx}  field={field!r}")
            failed += 1
            continue

        idx = races_df[mask].index[0]
        new_val = _coerce_value(field, value)
        old_val = races_df.at[idx, field]
        races_df.at[idx, field] = new_val
        if field == "surface":
            races_df.at[idx, "surface_source"] = "adjustment"
        if field in ("distance_m", "time_sec"):
            d = races_df.at[idx, "distance_m"]
            t = races_df.at[idx, "time_sec"]
            if d and t:
                races_df.at[idx, "pace_sec_per_mi"] = t / (d / 1609.344)
        if note is not None and not pd.isna(note) and str(note).strip():
            existing = races_df.at[idx, "note"] if "note" in races_df.columns else None
            if existing and not pd.isna(existing) and str(existing).strip():
                races_df.at[idx, "note"] = f"{existing} | {note}"
            else:
                races_df.at[idx, "note"] = note

        note_str = f"  ({note})" if (note is not None and not pd.isna(note) and str(note).strip()) else ""
        print(f"[adjustments] OK  {dt} seq={seq}  {field}: {old_val!r} → {new_val!r}{note_str}")
        applied += 1

    return races_df, applied, failed


def apply_adjustments(races_df, xlsx_path, sheet_name="changes"):
    """Legacy xlsx wrapper for apply_adjustments_from_df. Used only by freeze."""
    if not os.path.exists(xlsx_path):
        print(f"[adjustments] no file at {xlsx_path} — skipping")
        return races_df, 0, 0

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]
        print(f"[adjustments] no '{sheet_name}' sheet; falling back to '{ws.title}'")

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h:
            headers[str(h).strip().lower()] = c

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {k: ws.cell(row=r, column=c).value for k, c in headers.items()}
        rows.append(row)
    df = pd.DataFrame(rows)
    return apply_adjustments_from_df(races_df, df)


def ingest_additions_from_df(additions_df, source_label="snapshot:additions"):
    """Convert an additions DataFrame into race segment dicts.

    Columns required: date | distance_m | time_sec
    Optional:         surface | location | event | shoes | note
    """
    if additions_df is None or len(additions_df) == 0:
        return []
    required = {"date", "distance_m", "time_sec"}
    missing = required - set(additions_df.columns)
    if missing:
        print(f"[additions] missing columns: {missing} — skipping")
        return []

    rows = []
    for r_idx, row in additions_df.iterrows():
        date_v = row["date"]
        dist_v = row["distance_m"]
        time_v = row["time_sec"]
        if pd.isna(date_v) or pd.isna(dist_v) or pd.isna(time_v):
            continue

        if hasattr(date_v, "hour"):
            dt = date_v.date()
        elif hasattr(date_v, "year"):
            dt = date_v
        else:
            dt = pd.Timestamp(date_v).date()
        try:
            dist_m = int(float(dist_v))
        except (ValueError, TypeError):
            print(f"[additions] row {r_idx}: bad distance_m {dist_v!r} — skipping")
            continue
        if isinstance(time_v, (int, float)):
            time_sec = float(time_v)
        else:
            time_sec = parse_time_to_seconds(time_v)
        if time_sec is None:
            print(f"[additions] row {r_idx}: bad time_sec {time_v!r} — skipping")
            continue

        def _get(name):
            if name not in additions_df.columns:
                return None
            v = row.get(name)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return None
            s = str(v).strip()
            return s if s else None

        surface = _get("surface")
        location = _get("location")
        event = _get("event")
        shoes = _get("shoes")
        note = _get("note")

        miles_eq = dist_m / 1609.344
        pace = time_sec / miles_eq if miles_eq > 0 else None
        rows.append({
            "date": dt,
            "year": dt.year,
            "race_seq": 1,
            "fatigued": False,
            "distance_m": dist_m,
            "time_sec": time_sec,
            "pace_sec_per_mi": pace,
            "location": location,
            "temp_c": None,
            "weather": None,
            "conditions": None,
            "shoes": shoes,
            "surface": surface,
            "event": event,
            "note": note,
            "surface_source": "addition" if surface else None,
            "workout_raw": None,
            "source_file": source_label,
        })
        if note:
            print(f"[additions] added {dt} {dist_m}m {time_sec:.2f}s — {note}")
        else:
            print(f"[additions] added {dt} {dist_m}m {time_sec:.2f}s")
    return rows


def ingest_additions(xlsx_path, sheet_name="additions"):
    """Legacy xlsx wrapper for ingest_additions_from_df."""
    if not os.path.exists(xlsx_path):
        return []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h:
            headers[str(h).strip().lower()] = c
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {k: ws.cell(row=r, column=c).value for k, c in headers.items()}
        rows.append(row)
    df = pd.DataFrame(rows)
    return ingest_additions_from_df(df, source_label=os.path.basename(xlsx_path) + ":additions")


def append_additions(races_df, additions):
    """Append addition rows, skipping exact duplicates (same date+dist+time within 1%)."""
    if not additions:
        return races_df, 0, 0
    added = 0
    dup_warnings = 0
    new_rows = []
    races_dates = pd.to_datetime(races_df["date"]).dt.date
    for rr in additions:
        d = rr["date"]
        dist = rr["distance_m"]
        t = rr["time_sec"]
        mask = (races_dates == d) & (races_df["distance_m"] == dist)
        if mask.any():
            close = ((races_df.loc[mask, "time_sec"] - t).abs()
                     / races_df.loc[mask, "time_sec"].abs()) < 0.01
            if close.any():
                print(f"[additions] DUP WARN  {d} {dist}m {t:.2f}s already exists — skipping")
                dup_warnings += 1
                continue
        new_rows.append(rr)
        added += 1
    if new_rows:
        races_df = pd.concat([races_df, pd.DataFrame(new_rows)], ignore_index=True)
    races_df = races_df.sort_values("date").reset_index(drop=True)
    return races_df, added, dup_warnings


# ---------- autopop: location-based event/city-state derivation ----------

# Hardcoded rules: when the log's location matches (case-insensitive substring),
# set event to the canonical name and location to the city-state. If event is
# already set on the race, it's preserved — the rule only fills blanks.
HARDCODED_LOCATION_RULES = [
    {"match": "carnation", "event": "Run for the Pies", "city_state": "Carnation, WA"},
    {"match": "shoreline", "event": "Club Northwest All Comers", "city_state": "Shoreline, WA"},
    {"match": "boston",    "event": "Boston Marathon",           "city_state": "Boston, MA"},
]

# Event-name normalization: when a race has no log location but its event
# name matches one of these patterns, derive a location from the event.
#
# Matcher fields (first present wins per rule):
#   event_contains    — substring match (case-insensitive)
#   event_endswith    — suffix match (case-insensitive)
#   event_regex       — regex pattern (matched with re.IGNORECASE)
#
# Target fields:
#   infer_location    — chain to hardcoded rules + lookup (sets city_state
#                       and optionally event from the matched hardcoded rule)
#   city_state        — set location directly; event untouched
EVENT_NORMALIZATION_RULES = [
    {"event_contains": "All-Comers",          "infer_location": "shoreline"},
    {"event_contains": "All Comers",          "infer_location": "shoreline"},
    {"event_contains": "Run for the Pies",    "infer_location": "carnation"},
    {"event_contains": "Pies and Pints",      "infer_location": "carnation"},
    {"event_contains": "Winter Grand Prix",   "city_state":     "Seattle, WA"},
    # "Redmond" mentioned but NOT followed by '@' — catches "Eastlake @ Redmond"
    # (home meets hosted by Redmond HS) and "Redmond Invite" / "Redmond Derby"
    # style event names, while excluding "Redmond @ Eastlake" style away meets.
    {"event_regex":    r"\bRedmond\b(?!\s*@)", "city_state":    "Redmond, WA"},
]


def ingest_locations_from_df(locations_df):
    """Convert a locations DataFrame into the {log_location_lower: city_state} dict."""
    if locations_df is None or len(locations_df) == 0:
        return {}
    required = {"log_location", "city_state"}
    missing = required - set(locations_df.columns)
    if missing:
        print(f"[locations] missing columns: {missing} — skipping")
        return {}
    lookup = {}
    for _, row in locations_df.iterrows():
        loc = row["log_location"]
        city = row["city_state"]
        if pd.isna(loc) or pd.isna(city):
            continue
        loc = str(loc).strip().lower()
        city = str(city).strip()
        if loc and city:
            lookup[loc] = city
    return lookup


def ingest_locations(xlsx_path, sheet_name="locations"):
    """Legacy xlsx wrapper for ingest_locations_from_df."""
    if not os.path.exists(xlsx_path):
        return {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h:
            headers[str(h).strip().lower()] = c
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {k: ws.cell(row=r, column=c).value for k, c in headers.items()}
        rows.append(row)
    df = pd.DataFrame(rows)
    lookup = ingest_locations_from_df(df)
    print(f"[locations] loaded {len(lookup)} entries from '{sheet_name}' sheet")
    return lookup


def ingest_locations_csv(csv_path):
    """Fallback loader for location_lookup.csv when no locations sheet is
    available in the adjustments xlsx."""
    if not os.path.exists(csv_path):
        return {}
    lookup = {}
    with open(csv_path) as f:
        for r in csv.DictReader(f):
            loc = (r.get("log_location") or "").strip().lower()
            city = (r.get("city_state") or "").strip()
            if loc and city:
                lookup[loc] = city
    print(f"[locations] loaded {len(lookup)} entries from {csv_path}")
    return lookup


def apply_autopop(race_row, lookup, hardcoded_rules=None,
                  event_norm_rules=None):
    """Derive event/location for one race row. Mutates race_row in place.

    Precedence:
      1. Event-normalization: if log location is blank and event matches a
         known pattern, treat it as if the log location were the mapped
         value. This routes old All-Comers and Pies races to their
         hardcoded rules.
      2. Hardcoded rules: substring-match on (normalized) location. Sets
         city-state always; sets event only when event is blank.
      3. Lookup table: sets city-state from the lookup dict. Event untouched.
      4. Fall-through: returns 'flag' so caller knows nothing applied.

    Returns a status string: 'hardcoded' | 'lookup' | 'flag' | 'noop'.
    noop means event and city-state are both already non-trivially set
    (e.g. from an addition row with city-state event); nothing to do.
    """
    import csv  # noqa — already imported at module top, belt-and-braces
    hc = HARDCODED_LOCATION_RULES if hardcoded_rules is None else hardcoded_rules
    en = EVENT_NORMALIZATION_RULES if event_norm_rules is None else event_norm_rules

    event = race_row.get("event")
    event = "" if event is None or (isinstance(event, float) and pd.isna(event)) else str(event).strip()
    loc_raw = race_row.get("location")
    loc_raw = "" if loc_raw is None or (isinstance(loc_raw, float) and pd.isna(loc_raw)) else str(loc_raw).strip()
    loc_lower = loc_raw.lower()

    # If location is already in city-state format (contains comma) and event
    # is set, we've nothing to derive — leave alone.
    if "," in loc_raw and event:
        return "noop"

    # Step 1: event-based location normalization
    if not loc_lower and event:
        event_lower = event.lower()
        for rule in en:
            matched = False
            if "event_contains" in rule:
                matched = rule["event_contains"].lower() in event_lower
            elif "event_endswith" in rule:
                matched = event_lower.endswith(rule["event_endswith"].lower())
            elif "event_regex" in rule:
                matched = bool(re.search(rule["event_regex"], event, re.IGNORECASE))
            if not matched:
                continue
            # Direct city-state rule — set and return (event untouched)
            if "city_state" in rule:
                race_row["location"] = rule["city_state"]
                return "hardcoded"
            # Chain-to-lookup rule — set intermediate loc and fall through
            if "infer_location" in rule:
                loc_lower = rule["infer_location"]
                break

    # Step 2: hardcoded rules (substring match)
    for rule in hc:
        if rule["match"] in loc_lower:
            race_row["location"] = rule["city_state"]
            if not event:
                race_row["event"] = rule["event"]
            return "hardcoded"

    # Step 3: lookup table (exact match on lowercased key)
    if loc_lower in lookup:
        race_row["location"] = lookup[loc_lower]
        return "lookup"

    # Step 4: fall-through — nothing to do
    return "flag"

