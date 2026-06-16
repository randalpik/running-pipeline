"""Snapshot file format: reader, writer, and converters from other formats.

A snapshot concatenates seven CSV tables into one file, separated by
`# section: NAME [key=value]*` marker lines. Each section below its marker
is a well-formed CSV with a header row.

Sections:
  current_log year=YYYY   per-day rows for the in-progress current year
  changes                 race field overrides
  additions               manually-added race rows (primarily pre-2016)
  locations               log-location -> city-state lookup
  hills                   hill loop abbrev -> log_location (used by the
                          2016-17 location-synthesis pass during freeze)
  coordinates             city_state -> (lat, lon) overrides for cases
                          where Nominatim's lookup is wrong; applied on
                          top of the city_coords.csv cache so the override
                          is reproducible from source
  historical              city_state -> (min_hist, max_hist [, log_location])
                          date bounds for cities Max ran in. Two roles:
                          (a) surface remembered-but-unlogged cities on
                          the world map (e.g. Sapporo, JP); (b) override
                          city_state and (optionally) log_location on
                          non-race daily rows whose dates fall in the
                          range — replaces the legacy 2016-17
                          infer_2016_2017_location code path.

Missing sections return as empty DataFrames on read. Sections can appear in
any order.

Two conversion paths are supported:
  - from markdown (read_file_content dumps from the Drive MCP)
  - from xlsx     (OAuth Drive exports via drive_fetch)

CLI:
  python snapshot.py from-markdown --log-md PATH --log-year YYYY \\
                                   --adj-md PATH --out PATH
  python snapshot.py from-xlsx     --log-xlsx PATH --log-year YYYY \\
                                   --adj-xlsx PATH --out PATH
"""
import argparse
import csv
import io
import os
import re
import sys
from datetime import date, datetime
from typing import Any

import pandas as pd


# Output schema for the `current_log` section. Kept in sync with
# running_log_parser.CSV_LOG_COLUMNS (imported there as the authoritative
# list; duplicated here so this module stands alone).
CURRENT_LOG_COLUMNS = [
    "date", "sleep_cycles", "miles", "minutes", "temp_c", "weather",
    "workout_raw", "partners", "conditions", "wind", "wind_mph",
    "humidity_pct", "time_of_day", "shoes", "location", "weight_lbs",
]

# xlsx column positions per year era, mirroring running_log_parser.SCHEMAS.
# Kept local so snapshot.py doesn't reach into parser internals.
_XLSX_SCHEMAS = {
    # Standard layout 2017+ (2024/5 columns may or may not all be present).
    2017: {"month": 1, "day": 2, "sleep": 3, "miles": 4, "minutes": 5,
           "temp": 7, "weather": 8, "workout": 9, "partners": 10,
           "conditions": None, "wind": None, "time_of_day": None,
           "shoes": None, "location": None, "weight": None,
           "sleep_unit": "hours"},
    2018: {"month": 2, "day": 3, "sleep": 4, "miles": 5, "minutes": 6,
           "temp": 7, "weather": 8, "workout": 9, "partners": 10,
           "conditions": 11, "wind": None, "time_of_day": 12,
           "shoes": None, "location": 13, "weight": None,
           "sleep_unit": "cycles"},
    2020: {"month": 2, "day": 3, "sleep": 4, "miles": 5, "minutes": 6,
           "temp": 7, "weather": 8, "workout": 9, "partners": 10,
           "conditions": 11, "wind": None, "time_of_day": 12,
           "shoes": 13, "location": 14, "weight": None,
           "sleep_unit": "cycles"},
    2022: {"month": 2, "day": 3, "sleep": 4, "miles": 5, "minutes": 6,
           "temp": 7, "weather": 8, "workout": 9, "partners": 10,
           "conditions": 11, "wind": None, "time_of_day": 12,
           "shoes": 13, "location": 14, "weight": 15,
           "sleep_unit": "cycles"},
    2025: {"month": 2, "day": 3, "sleep": 4, "miles": 5, "minutes": 6,
           "temp": 7, "weather": 8, "workout": 9, "partners": 10,
           "conditions": 11, "wind": 12, "time_of_day": 13,
           "shoes": 14, "location": 15, "weight": 16,
           "sleep_unit": "cycles"},
    # 2026 reorg: sleep/temp/wind/time_of_day dropped from the hand log (now
    # sourced from watch data); remaining columns reordered. None-valued fields
    # pass through blank; the watch override refills temp/time_of_day/wind_mph.
    2026: {"month": 2, "day": 3, "sleep": None, "miles": 4, "minutes": 5,
           "temp": None, "weather": 6, "workout": 8, "partners": 11,
           "conditions": 7, "wind": None, "time_of_day": None,
           "shoes": 9, "location": 10, "weight": 12,
           "sleep_unit": "cycles"},
}

_YEAR_TO_SCHEMA = {
    2017: 2017, 2018: 2018, 2019: 2018,
    2020: 2020, 2021: 2020,
    2022: 2022, 2023: 2022, 2024: 2022,
    2025: 2025, 2026: 2026, 2027: 2026,
}


# ---------- snapshot read / find ----------

_SECTION_HEADER_RE = re.compile(r'^\s*#\s*section\s*:\s*(\S+)(.*)$')


def _parse_section_header(line):
    m = _SECTION_HEADER_RE.match(line)
    if not m:
        return None
    name = m.group(1).strip()
    meta = {}
    for kv in re.finditer(r'(\w+)=(\S+)', m.group(2) or ''):
        meta[kv.group(1)] = kv.group(2)
    return name, meta


def read_snapshot(path):
    """Parse a snapshot file. Returns (sections_dict, meta_dict).

    sections_dict maps section name -> DataFrame (possibly empty).
    meta_dict maps section name -> dict of `key=value` metadata from the marker.
    """
    with open(path) as f:
        lines = f.readlines()

    sections = {}
    metas = {}
    current_name = None
    current_meta = None
    current_buf = []

    def _flush():
        if current_name is None:
            return
        body = ''.join(current_buf).strip()
        if body:
            try:
                df = pd.read_csv(io.StringIO(body))
            except pd.errors.EmptyDataError:
                df = pd.DataFrame()
        else:
            df = pd.DataFrame()
        sections[current_name] = df
        metas[current_name] = current_meta or {}

    for line in lines:
        hdr = _parse_section_header(line)
        if hdr is not None:
            _flush()
            current_name, current_meta = hdr
            current_buf = []
        else:
            if current_name is not None:
                current_buf.append(line)
    _flush()
    return sections, metas


def find_snapshot(candidate_paths):
    """Return the first existing path from `candidate_paths`, or None."""
    for p in candidate_paths:
        if p and os.path.exists(p):
            return p
    return None


# ---------- snapshot write ----------

def _write_section(buf, name, header, rows, meta=""):
    """Write one section block to `buf`. `rows` is an iterable of row lists."""
    buf.write(f"# section: {name}{meta}\n")
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
    w.writerow(header)
    for row in rows:
        w.writerow(["" if (c is None or (isinstance(c, float) and pd.isna(c)))
                    else c for c in row])
    buf.write("\n")


def _df_to_rows(df, columns):
    """Yield row lists from a DataFrame, one per record, aligned to `columns`."""
    if df is None or df.empty:
        return
    for _, r in df.iterrows():
        yield [r.get(c) for c in columns]


def write_snapshot(path, *, current_year, current_log_df,
                   changes_df, additions_df, locations_df, hills_df=None,
                   coordinates_df=None, historical_df=None):
    """Write a snapshot file at `path`."""
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    buf = io.StringIO()

    _write_section(buf, "current_log", CURRENT_LOG_COLUMNS,
                   _df_to_rows(current_log_df, CURRENT_LOG_COLUMNS),
                   meta=f" year={current_year}")

    changes_cols = ["date", "race_seq", "field", "value", "note"]
    _write_section(buf, "changes", changes_cols,
                   _df_to_rows(changes_df, changes_cols))

    additions_cols = ["date", "distance_m", "time_sec", "surface",
                      "location", "event"]
    _write_section(buf, "additions", additions_cols,
                   _df_to_rows(additions_df, additions_cols))

    # Preserve any extra metadata columns present in the locations sheet
    # (display_name, elev_per_mile, altitude, terrain_type, plus anything
    # added later). Core columns come first; extras appended in source order.
    core_locations_cols = ["log_location", "city_state"]
    extra_locations_cols = (
        [c for c in locations_df.columns if c not in core_locations_cols]
        if locations_df is not None and not locations_df.empty
        else []
    )
    locations_cols = core_locations_cols + extra_locations_cols
    _write_section(buf, "locations", locations_cols,
                   _df_to_rows(locations_df, locations_cols))

    # hills: abbrev -> location with optional metadata (type, distance_m,
    # elev_*). Same pattern as locations — core columns first, extras kept.
    core_hills_cols = ["abbrev", "location"]
    extra_hills_cols = (
        [c for c in hills_df.columns if c not in core_hills_cols]
        if hills_df is not None and not hills_df.empty
        else []
    )
    hills_cols = core_hills_cols + extra_hills_cols
    _write_section(buf, "hills", hills_cols,
                   _df_to_rows(hills_df, hills_cols))

    # coordinates: city_state -> (latitude, longitude) overrides applied on
    # top of Nominatim's cached results. Empty section is fine; written so
    # snapshots always have a stable shape.
    coordinates_cols = ["city_state", "latitude", "longitude"]
    _write_section(buf, "coordinates", coordinates_cols,
                   _df_to_rows(coordinates_df, coordinates_cols))

    # historical: city_state -> (min_hist, max_hist [, log_location]) for
    # remembered-but-unlogged cities AND for date-range-driven overrides
    # of non-race daily rows. Empty section keeps the on-disk shape stable.
    # log_location is optional — when set, the daily-row override also
    # writes it onto `location` so route-bin signal (e.g. 'nashville',
    # 'education hill') survives for recovery-analysis groupings.
    historical_cols = ["city_state", "min_hist", "max_hist", "log_location"]
    _write_section(buf, "historical", historical_cols,
                   _df_to_rows(historical_df, historical_cols))

    with open(path, "w") as f:
        f.write(buf.getvalue())
    return len(buf.getvalue())


# ---------- markdown → snapshot ----------

def _unescape_md(s):
    if s is None:
        return None
    s = s.strip()
    if s == "":
        return None
    return re.sub(r"\\([\-\*_\[\]\\\|])", r"\1", s)


def _split_md_tables(md_text):
    """Yield lists of rows per table block in the markdown."""
    lines = md_text.splitlines()
    i = 0
    while i < len(lines):
        while i < len(lines) and ":-:" not in lines[i]:
            i += 1
        if i >= len(lines):
            return
        header_cells = [c.strip() for c in lines[i - 1].strip("|").split("|")]
        rows = [header_cells]
        j = i + 1
        while j < len(lines) and lines[j].strip().startswith("|"):
            cells = [_unescape_md(c) for c in lines[j].strip("|").split("|")]
            rows.append(cells)
            j += 1
        yield rows
        i = j
        while i < len(lines) and lines[i].strip() == "":
            i += 1


def current_log_df_from_markdown(md_path, year):
    """Parse a markdown log dump into the current_log DataFrame."""
    with open(md_path) as f:
        md = f.read()
    tables = list(_split_md_tables(md))
    if not tables:
        raise ValueError(f"no tables in {md_path}")
    table = tables[0]

    out_rows = []
    for row in table[1:]:
        if len(row) < 16:
            row = row + [None] * (16 - len(row))
        date_str = row[2]
        if not date_str:
            continue
        try:
            mo, dy = date_str.split("/")
            dt = date(year, int(mo), int(dy))
        except (ValueError, AttributeError):
            continue
        # Skip rows with no logged activity
        if not row[4] and not row[8]:
            continue
        out_rows.append({
            "date": dt.isoformat(),
            "sleep_cycles": row[3],
            "miles": row[4],
            "minutes": row[5],
            "temp_c": row[6],
            "weather": row[7],
            "workout_raw": row[8],
            "partners": row[9],
            "conditions": row[10],
            "wind": row[11],
            "time_of_day": row[12],
            "shoes": row[13],
            "location": row[14],
            "weight_lbs": row[15],
        })
    return pd.DataFrame(out_rows, columns=CURRENT_LOG_COLUMNS)


def adjustments_dfs_from_markdown(md_path):
    """Parse a markdown adjustments dump into (changes, additions, locations,
    hills, coordinates, historical).

    The hills, coordinates, and historical tables are optional for backward
    compatibility with older dumps. When absent they're returned as empty
    DataFrames.
    """
    with open(md_path) as f:
        md = f.read()
    tables = list(_split_md_tables(md))
    if len(tables) < 3 or len(tables) > 6:
        raise ValueError(f"expected 3-6 adjustment tables, got {len(tables)}")

    def _classify(t):
        hdr = "|".join((_unescape_md(h) or "").lower() for h in t[0])
        if "race_seq" in hdr:
            return "changes"
        if "distance_m" in hdr:
            return "additions"
        if "log_location" in hdr:
            return "locations"
        if "abbrev" in hdr:
            return "hills"
        if "min_hist" in hdr or "max_hist" in hdr:
            return "historical"
        if "latitude" in hdr or "longitude" in hdr:
            return "coordinates"
        return None

    by_name = {}
    for t in tables:
        name = _classify(t)
        if name is None:
            raise ValueError(f"couldn't classify table with header {t[0]!r}")
        by_name[name] = t
    for name in ("changes", "additions", "locations"):
        if name not in by_name:
            raise ValueError(f"missing adjustment section: {name}")

    def _to_df(t):
        header = [_unescape_md(h) for h in t[0]]
        data = [dict(zip(header, row + [None] * (len(header) - len(row))))
                for row in t[1:]]
        return pd.DataFrame(data, columns=header)

    hills_df = _to_df(by_name["hills"]) if "hills" in by_name else pd.DataFrame()
    coords_df = (_to_df(by_name["coordinates"])
                 if "coordinates" in by_name else pd.DataFrame())
    historical_df = (_to_df(by_name["historical"])
                     if "historical" in by_name else pd.DataFrame())
    return (_to_df(by_name["changes"]),
            _to_df(by_name["additions"]),
            _to_df(by_name["locations"]),
            hills_df,
            coords_df,
            historical_df)


def snapshot_from_markdown(log_md_path, log_year, adj_md_path, out_path):
    """High-level: markdown files → snapshot CSV at `out_path`."""
    log_df = current_log_df_from_markdown(log_md_path, log_year)
    (changes_df, additions_df, locations_df, hills_df, coordinates_df,
     historical_df) = adjustments_dfs_from_markdown(adj_md_path)
    size = write_snapshot(out_path,
                          current_year=log_year,
                          current_log_df=log_df,
                          changes_df=changes_df,
                          additions_df=additions_df,
                          locations_df=locations_df,
                          hills_df=hills_df,
                          coordinates_df=coordinates_df,
                          historical_df=historical_df)
    print(f"[snapshot] wrote {out_path}  ({size} bytes)  "
          f"log={len(log_df)}, changes={len(changes_df)}, "
          f"additions={len(additions_df)}, locations={len(locations_df)}, "
          f"hills={len(hills_df)}, coordinates={len(coordinates_df)}, "
          f"historical={len(historical_df)}")


# ---------- xlsx → snapshot ----------

def _xlsx_cell(ws, row, col) -> Any:
    if col is None:
        return None
    v = ws.cell(row=row, column=col).value
    return None if v in (None, "") else v


def current_log_df_from_xlsx(xlsx_path, year, sheet_name=None):
    """Parse a Running Log YYYY xlsx into the current_log DataFrame.

    Accepts the standard 2017+ layout (data rows begin at row 3, day = row-2).
    """
    import openpyxl

    schema_key = _YEAR_TO_SCHEMA.get(year)
    if schema_key is None:
        raise ValueError(f"no xlsx schema for year {year}")
    sch = _XLSX_SCHEMAS[schema_key]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Prefer 'Data' if present (2024+), else first sheet
        ws = wb["Data"] if "Data" in wb.sheetnames else wb.worksheets[0]

    from datetime import timedelta
    max_doy = 366 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 365
    out_rows = []
    for r in range(3, 3 + max_doy):
        miles_v = _xlsx_cell(ws, r, sch["miles"])
        minutes_v = _xlsx_cell(ws, r, sch["minutes"])
        workout_v = _xlsx_cell(ws, r, sch["workout"])
        sleep_v = _xlsx_cell(ws, r, sch["sleep"])
        if miles_v is None and minutes_v is None and workout_v is None and sleep_v is None:
            continue
        doy = r - 2
        dt = date(year, 1, 1) + timedelta(days=doy - 1)

        sleep_cycles = sleep_v
        if sleep_v is not None and sch["sleep_unit"] == "hours":
            try:
                sleep_cycles = float(sleep_v) / 1.5
            except (ValueError, TypeError):
                sleep_cycles = sleep_v

        out_rows.append({
            "date": dt.isoformat(),
            "sleep_cycles": sleep_cycles,
            "miles": miles_v,
            "minutes": minutes_v,
            "temp_c": _xlsx_cell(ws, r, sch["temp"]),
            "weather": _xlsx_cell(ws, r, sch["weather"]),
            "workout_raw": workout_v,
            "partners": _xlsx_cell(ws, r, sch["partners"]),
            "conditions": _xlsx_cell(ws, r, sch["conditions"]),
            "wind": _xlsx_cell(ws, r, sch["wind"]),
            "time_of_day": _xlsx_cell(ws, r, sch["time_of_day"]),
            "shoes": _xlsx_cell(ws, r, sch["shoes"]),
            "location": _xlsx_cell(ws, r, sch["location"]),
            "weight_lbs": _xlsx_cell(ws, r, sch["weight"]),
        })
    return pd.DataFrame(out_rows, columns=CURRENT_LOG_COLUMNS)


def adjustments_dfs_from_xlsx(xlsx_path):
    """Parse the adjustments xlsx into (changes, additions, locations, hills,
    coordinates, historical) DataFrames.

    The hills, coordinates, and historical sheets are optional; absence
    yields an empty DataFrame.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    def _sheet_to_df(sheet_name, expected_cols):
        if sheet_name not in wb.sheetnames:
            return pd.DataFrame(columns=expected_cols)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return pd.DataFrame(columns=expected_cols)
        headers = [str(h).strip() if h else "" for h in rows[0]]
        data = []
        for r in rows[1:]:
            rec = dict(zip(headers, r))
            if all(v in (None, "") for v in rec.values()):
                continue
            # Normalize date-like values to ISO string for CSV consistency.
            # Applies to `date` (changes/additions) plus the historical
            # date-bound columns.
            for date_col in ("date", "min_hist", "max_hist"):
                if date_col in rec and rec[date_col] not in (None, ""):
                    d = rec[date_col]
                    if isinstance(d, datetime):
                        rec[date_col] = d.date().isoformat()
                    elif isinstance(d, date):
                        rec[date_col] = d.isoformat()
            data.append(rec)
        return pd.DataFrame(data, columns=expected_cols + [c for c in headers if c not in expected_cols])

    changes = _sheet_to_df("changes", ["date", "race_seq", "field", "value", "note"])
    additions = _sheet_to_df("additions", ["date", "distance_m", "time_sec",
                                           "surface", "location", "event"])
    locations = _sheet_to_df("locations", ["log_location", "city_state"])
    hills = _sheet_to_df("hills", ["abbrev", "location"])
    coordinates = _sheet_to_df("coordinates",
                               ["city_state", "latitude", "longitude"])
    historical = _sheet_to_df("historical",
                              ["city_state", "min_hist", "max_hist",
                               "log_location"])
    return changes, additions, locations, hills, coordinates, historical


def snapshot_from_xlsx(log_xlsx_path, log_year, adj_xlsx_path, out_path):
    """High-level: xlsx files → snapshot CSV at `out_path`."""
    log_df = current_log_df_from_xlsx(log_xlsx_path, log_year)
    (changes_df, additions_df, locations_df, hills_df, coordinates_df,
     historical_df) = adjustments_dfs_from_xlsx(adj_xlsx_path)
    size = write_snapshot(out_path,
                          current_year=log_year,
                          current_log_df=log_df,
                          changes_df=changes_df,
                          additions_df=additions_df,
                          locations_df=locations_df,
                          hills_df=hills_df,
                          coordinates_df=coordinates_df,
                          historical_df=historical_df)
    print(f"[snapshot] wrote {out_path}  ({size} bytes)  "
          f"log={len(log_df)}, changes={len(changes_df)}, "
          f"additions={len(additions_df)}, locations={len(locations_df)}, "
          f"hills={len(hills_df)}, coordinates={len(coordinates_df)}, "
          f"historical={len(historical_df)}")


# ---------- CLI ----------

def main():
    p = argparse.ArgumentParser(description=(__doc__ or "").split("\n\n")[0])
    sub = p.add_subparsers(dest="cmd", required=True)

    md = sub.add_parser("from-markdown",
                        help="Build a snapshot from read_file_content markdown dumps")
    md.add_argument("--log-md", required=True)
    md.add_argument("--log-year", type=int, required=True)
    md.add_argument("--adj-md", required=True)
    md.add_argument("--out", required=True)

    xl = sub.add_parser("from-xlsx",
                        help="Build a snapshot from Drive-exported xlsx files")
    xl.add_argument("--log-xlsx", required=True)
    xl.add_argument("--log-year", type=int, required=True)
    xl.add_argument("--adj-xlsx", required=True)
    xl.add_argument("--out", required=True)

    args = p.parse_args()
    if args.cmd == "from-markdown":
        snapshot_from_markdown(args.log_md, args.log_year, args.adj_md, args.out)
    elif args.cmd == "from-xlsx":
        snapshot_from_xlsx(args.log_xlsx, args.log_year, args.adj_xlsx, args.out)


if __name__ == "__main__":
    main()
